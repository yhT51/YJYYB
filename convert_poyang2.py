"""
将鄱阳工单数据转换为和革命伤残军人休养院运维报告相同的格式
"""

from openpyxl import load_workbook

# 读取两个文件
poyang_file = r"D:\XLX\XLXWJ\鄱阳工单数据_1773391976677.xlsx"
soldier_file = r"D:\XLX\XLXWJ\革命伤残军人休养院云净血透系统运维报告.docx"

print("读取鄱阳工单数据...")
try:
    wb = load_workbook(poyang_file)
    ws = wb.active

    print(f"成功读取")
    print(f"工作表名称: {ws.title}")
    print(f"数据行数: {ws.max_row}")
    print(f"数据列数: {ws.max_column}")

    print("\n前5行数据:")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)

    print("\n列名:")
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for i, col in enumerate(header_row):
        print(f"列{i+1}: {col}")

except Exception as e:
    print(f"读取失败: {e}")
    input("按回车退出...")
    exit(1)

print("\n" + "="*50)
print("革命伤残军人休养院运维报告文件分析")
print("注意: 该文件是 .docx 格式")
print("需要提取文档结构和格式作为模板")
print("="*50)
