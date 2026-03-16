from openpyxl import load_workbook
from datetime import datetime

# 读取Excel数据
wb = load_workbook('D:\\XLX\\XLXWJ\\鄱阳工单数据_1773391976677.xlsx')
ws = wb.active

# 先收集所有数据
lines = []

# 跳过标题行，从第2行开始
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[15]:  # 完成时间（第16列，索引15）
        try:
            dt = row[15]
            if isinstance(dt, datetime):
                date_str = dt.strftime('%Y年%m月%d日')
            else:
                date_str = str(dt)[:10].replace('-', '年').replace('-', '月') + '日'
            
            title = row[7] if row[7] else ''  # 工单标题（H列，索引7）
            
            if title:
                line = f'{date_str}，{title}'
                lines.append(line)
        except:
            continue

# 创建HTML文档
html_content = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>鄱阳工单数据报告</title>
</head>
<body>
    <h1>鄱阳工单数据报告</h1>
    <p><br></p>
'''

for line in lines:
    html_content += f'    <p>{line}</p>\n'

html_content += '''</body>
</html>'''

# 保存HTML文件
output_path = 'D:\\XLX\\XLXWJ\\鄱阳工单数据报告.html'
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(html_content)
print(f'报告已生成：{output_path}')
